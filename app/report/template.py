from __future__ import annotations
from pathlib import Path
import time, json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..io.loader import write_table
from ..core.profile import profile_dataframe
from ..autoexcel.engines_fallback import create_pivot_from_df, add_chart

THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _fit(ws):
    """워크시트 열 너비 자동 조정"""
    ws.freeze_panes = "A2"
    for i, c in enumerate(ws[1], 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(c.value or "")) + 4)

def _sheet_cover(wb, title: str, period: str, owner: str):
    """표지 시트 생성"""
    ws = wb.create_sheet("00_표지")
    ws["A1"] = title
    ws["A1"].font = Font(size=20, bold=True)
    ws["A3"] = "기간"
    ws["B3"] = period
    ws["A4"] = "작성자"
    ws["B4"] = owner
    ws["A6"] = "생성시각"
    ws["B6"] = time.strftime("%Y-%m-%d %H:%M:%S")
    
    for r in (3, 4, 6):
        ws[f"A{r}"].font = Font(bold=True)
    
    return ws

def _sheet_kpi(wb, df: pd.DataFrame, t_clean: float|None = None):
    """KPI 시트 생성"""
    ws = wb.create_sheet("01_KPI")
    prof = profile_dataframe(df)
    
    # 간단 KPI
    rows, cols = df.shape
    miss = float(df.isna().mean().mean())
    dup_rate = float((rows - len(df.drop_duplicates())) / max(1, rows))
    
    kpis = [
        ("행 수", rows),
        ("열 수", cols),
        ("결측률(평균)", round(miss, 4)),
        ("중복률(전체기준)", round(dup_rate, 4)),
    ]
    
    if t_clean is not None:
        kpis.append(("전처리 시간(s)", round(t_clean, 3)))
    
    ws.append(["지표", "값"])
    for k, v in kpis:
        ws.append([k, v])
    
    # 스타일 적용
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
    
    for r in range(1, ws.max_row + 1):
        for c in ("A", "B"):
            ws[f"{c}{r}"].border = BORDER
    
    _fit(ws)
    
    # 전체 프로파일 JSON은 숨김 시트로 저장(필요시)
    ws2 = wb.create_sheet("99_profile_json")
    ws2.sheet_state = "hidden"
    ws2["A1"] = json.dumps(prof, ensure_ascii=False, default=str)
    
    return ws

def _sheet_sample(wb, df: pd.DataFrame, n=50):
    """원본 샘플 시트 생성"""
    ws = wb.create_sheet("02_원본샘플")
    ws.append(list(df.columns))
    
    for _, row in df.head(n).iterrows():
        ws.append(list(row.values))
    
    # 스타일 적용
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF2FF")
    
    _fit(ws)
    return ws

def _apply_number_formats(path_xlsx: str, sheet: str, value_cols: list[str], fmt="₩#,##0"):
    """숫자/통화 서식 적용"""
    wb = load_workbook(path_xlsx)
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    
    idx = [header.index(v) + 1 for v in value_cols if v in header]
    
    for col_i in idx:
        col = get_column_letter(col_i)
        for r in range(2, ws.max_row + 1):
            ws[f"{col}{r}"].number_format = fmt
    
    wb.save(path_xlsx)

def build_report(df: pd.DataFrame, out_path: str,
                 title="월별 카테고리 매출 보고서", period="최근 기간", owner="Excel Copilot",
                 rows=("월", "카테고리"), values=(("금액", "sum"),), filters=None,
                 chart_type="bar", apply_currency_format=True):
    """완전한 보고서 생성"""
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    
    # 0) 베이스 워크북
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "피벗_결과(초기화)"
    wb.save(out_path)  # 초기 세이브
    
    # 1) 표지/KPI/원본샘플
    wb = load_workbook(out_path)
    _sheet_cover(wb, title, period, owner)
    _sheet_kpi(wb, df)
    _sheet_sample(wb, df)
    wb.save(out_path)
    
    # 2) 피벗 + 차트 (기존 엔진 재사용)
    shape = create_pivot_from_df(df, out_path, "03_피벗", list(rows), list(values), filters=filters or {})
    add_chart(out_path, "03_피벗", title=title, chart_type=chart_type)
    
    # 3) 값 서식(통화)
    if apply_currency_format:
        val_cols = []
        # 피벗 결과 헤더에서 "금액" 포함 열을 자동 식별
        wb2 = load_workbook(out_path)
        hdr = [c.value for c in wb2["03_피벗"][1]]
        for h in hdr:
            if h and ("금액" in str(h) or "sum" in str(h).lower()):
                val_cols.append(str(h))
        wb2.close()
        
        if val_cols:
            _apply_number_formats(out_path, "03_피벗", val_cols)
    
    return {"out": out_path, "pivot_shape": shape}
