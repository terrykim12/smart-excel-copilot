from __future__ import annotations
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def _col_index_by_header(ws, name_contains: str):
    hdr = [c.value for c in ws[1]]
    idx = [i+1 for i,h in enumerate(hdr) if h and name_contains in str(h)]
    return idx, hdr

def add_mom_ytd_columns(xlsx_path: str, sheet: str = "03_피벗"):
    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        # 피벗 시트 추정
        cand = [s for s in wb.sheetnames if "피벗" in s]
        if not cand: raise ValueError("피벗 시트를 찾을 수 없습니다")
        sheet = cand[0]
    ws = wb[sheet]
    rows = ws.max_row; cols = ws.max_column
    if rows < 3 or cols < 3: 
        wb.save(xlsx_path); return

    # 가정: A=월(YYYY-MM 또는 날짜), B=범주(카테고리), C..=값 열들
    val_start = 3
    for c in range(val_start, cols+1):
        val_header = ws.cell(row=1, column=c).value
        if not val_header: 
            continue
        # 새 헤더
        mom_col = cols + 1; ytd_col = cols + 2
        ws.cell(row=1, column=mom_col).value = f"{val_header}_MoM"
        ws.cell(row=1, column=ytd_col).value = f"{val_header}_YTD"
        # 각 행에 수식 입력
        for r in range(2, rows+1):
            # 전월대비 = 같은 카테고리이고 바로 위 행이면 (현재-이전)/이전
            ws.cell(row=r, column=mom_col).value = (
                f"=IF($B{r}=$B{r-1}, IFERROR(({get_column_letter(c)}{r}-{get_column_letter(c)}{r-1})/{get_column_letter(c)}{r-1},0), 0)"
            )
            # 누계(YTD) = 같은 카테고리 및 월<=현재월 조건 SUMIFS
            ws.cell(row=r, column=ytd_col).value = (
                f"=SUMIFS({get_column_letter(c)}:{get_column_letter(c)}, $B:$B, $B{r}, $A:$A, \"<=\"&$A{r})"
            )
        cols += 2  # 다음 값열을 위해 확장

    wb.save(xlsx_path)
