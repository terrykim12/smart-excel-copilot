#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd

class OpenPyxlExcel:
    """openpyxl을 사용한 Excel 처리 엔진"""
    
    def __init__(self):
        self.workbook = None
        self.active_sheet = None
    
    def open(self, path):
        """Excel 파일을 엽니다."""
        try:
            if os.path.exists(path):
                self.workbook = load_workbook(path)
                self.active_sheet = self.workbook.active
                print(f"[debug] 파일 열기 성공: {path}")
            else:
                # 파일이 없으면 새 워크북 생성
                self.workbook = Workbook()
                self.active_sheet = self.workbook.active
                print(f"[debug] 새 워크북 생성: {path}")
            return self.workbook
        except Exception as e:
            print(f"[debug] 파일 열기 실패: {e}")
            # 새 워크북 생성
            self.workbook = Workbook()
            self.active_sheet = self.workbook.active
            return self.workbook
    
    def apply_number_formats(self, ws, value_cols, fmt="₩#,##0"):
        """숫자/통화 서식을 일괄 적용합니다."""
        try:
            # 헤더에서 값 열 찾기
            headers = [cell.value for cell in ws[1]]
            for i, header in enumerate(headers, start=1):
                if header and any(col in str(header) for col in value_cols):
                    col_letter = get_column_letter(i)
                    print(f"[debug] 서식 적용: {col_letter}열 ({header})")
                    
                    # 헤더 제외하고 데이터 행에 서식 적용
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f"{col_letter}{row}"]
                        if cell.value is not None:
                            cell.number_format = fmt
                    
                    print(f"[debug] {col_letter}열 서식 적용 완료")
            
            return True
        except Exception as e:
            print(f"[debug] 서식 적용 실패: {e}")
            return False
    
    def create_pivot_chart(self, src_sheet, tgt_sheet, rows, values, chart_type="bar", value_format="₩#,##0"):
        """피벗 테이블과 차트를 생성합니다."""
        try:
            # 소스 시트 선택
            if src_sheet in self.workbook.sheetnames:
                ws = self.workbook[src_sheet]
            else:
                ws = self.workbook.active
            
            print(f"[debug] 소스 시트: {ws.title}")
            
            # 데이터를 pandas DataFrame으로 변환
            data = []
            headers = []
            
            for row in ws.iter_rows(values_only=True):
                if not headers:
                    headers = row
                    continue
                if any(cell is not None for cell in row):
                    data.append(row)
            
            if not data:
                raise ValueError("데이터가 없습니다")
            
            df = pd.DataFrame(data, columns=headers)
            print(f"[debug] 데이터 로드 완료: {len(df)}행 x {len(df.columns)}열")
            
            # 피벗 테이블 생성
            pivot_data = df.groupby(rows)[values].sum().reset_index()
            print(f"[debug] 피벗 데이터: {pivot_data}")
            
            # 대상 시트 생성 또는 선택
            if tgt_sheet in self.workbook.sheetnames:
                target_ws = self.workbook[tgt_sheet]
            else:
                target_ws = self.workbook.create_sheet(tgt_sheet)
            
            # 피벗 데이터를 시트에 작성
            target_ws['A1'] = f"{rows[0]}별 {values[0]} 합계"
            
            # 헤더 작성
            for col, header in enumerate(pivot_data.columns, 1):
                target_ws.cell(row=2, column=col, value=header)
            
            # 데이터 작성
            for row_idx, (_, row_data) in enumerate(pivot_data.iterrows(), 3):
                for col_idx, value in enumerate(row_data, 1):
                    target_ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 숫자 서식 적용
            if value_format:
                self.apply_number_formats(target_ws, values, value_format)
            
            # 차트 생성을 위해 대상 시트를 활성 시트로 설정
            self.active_sheet = target_ws
            
            # 차트 생성
            chart = self._create_chart(chart_type, pivot_data, rows[0], values[0], value_format)
            
            # 차트를 시트에 추가
            target_ws.add_chart(chart, "D2")
            
            print(f"[debug] 피벗 차트 생성 완료: {tgt_sheet}")
            return True
            
        except Exception as e:
            print(f"[debug] 피벗 차트 생성 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_chart(self, chart_type, data, x_label, y_label, value_format="₩#,##0"):
        """차트 타입에 따라 적절한 차트를 생성합니다."""
        if chart_type.lower() == "bar":
            chart = BarChart()
        elif chart_type.lower() == "line":
            chart = LineChart()
        elif chart_type.lower() == "pie":
            chart = PieChart()
        else:
            chart = BarChart()  # 기본값
        
        chart.title = f"{x_label}별 {y_label} 차트"
        chart.x_axis.title = x_label
        chart.y_axis.title = y_label
        
        # 차트 값 축에 통화 포맷 적용
        if value_format and hasattr(chart.y_axis, 'number_format'):
            chart.y_axis.number_format = value_format
            print(f"[debug] 차트 값 축 서식 적용: {value_format}")
        
        # 데이터 범위 설정 - 워크시트의 실제 셀 범위를 사용
        # 피벗 데이터가 이미 워크시트에 작성되어 있으므로 해당 범위 참조
        max_row = len(data) + 2  # 헤더 포함
        
        if chart_type.lower() == "pie":
            # 파이 차트는 값과 라벨만 필요
            data_ref = Reference(self.active_sheet, min_col=2, min_row=2, max_row=max_row)
            cats_ref = Reference(self.active_sheet, min_col=1, min_row=3, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
        else:
            # 막대/선 차트
            data_ref = Reference(self.active_sheet, min_col=2, min_row=2, max_row=max_row)
            cats_ref = Reference(self.active_sheet, min_col=1, min_row=3, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
        
        return chart
    
    def save_as(self, path):
        """워크북을 저장합니다."""
        try:
            self.workbook.save(path)
            print(f"[debug] 파일 저장 완료: {path}")
            return True
        except Exception as e:
            print(f"[debug] 파일 저장 실패: {e}")
            return False
    
    def close(self):
        """워크북을 닫습니다."""
        try:
            if self.workbook:
                self.workbook.close()
                print("[debug] 워크북 닫기 완료")
        except Exception as e:
            print(f"[debug] 워크북 닫기 실패: {e}")

def create_pivot_chart_openpyxl(path_in, path_out, src_sheet, tgt_sheet, rows, values, chart_type="bar", value_format="₩#,##0"):
    """openpyxl을 사용하여 피벗 차트를 생성하는 함수"""
    
    # 절대 경로로 변환
    abs_path_in = os.path.abspath(path_in)
    abs_path_out = os.path.abspath(path_out)
    
    print(f"[debug] 입력 파일 절대 경로: {abs_path_in}")
    print(f"[debug] 출력 파일 절대 경로: {abs_path_out}")
    
    # 파일 존재 확인
    if not os.path.exists(abs_path_in):
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {abs_path_in}")
    
    # 출력 디렉토리 생성
    os.makedirs(os.path.dirname(abs_path_out), exist_ok=True)
    
    engine = OpenPyxlExcel()
    try:
        # 파일 열기
        wb = engine.open(abs_path_in)
        
        # 피벗 차트 생성
        success = engine.create_pivot_chart(src_sheet, tgt_sheet, rows, values, chart_type, value_format)
        
        if success:
            # 파일 저장
            engine.save_as(abs_path_out)
            print(f"✅ 피벗 차트 생성 및 저장 완료: {abs_path_out}")
        else:
            print("❌ 피벗 차트 생성 실패")
            
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        engine.close()
    
    return success
