
from __future__ import annotations
import sys, os, re, json, time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

try:
    if sys.platform == "win32":
        import win32com.client as win32  # type: ignore
    else:
        win32 = None
except Exception:
    win32 = None

A1_RE = re.compile(r"^(?:(?P<sheet>[^!]+)!)?(?P<start_col>[A-Za-z]+)(?P<start_row>\d+):(?P<end_col>[A-Za-z]+)(?P<end_row>\d+)$")

def a1_to_range(ws, a1: str):
    m = A1_RE.match(a1)
    if not m:
        raise ValueError(f"Invalid A1 range: {a1}")
    srow = int(m.group("start_row")); erow = int(m.group("end_row"))
    scol = m.group("start_col").upper(); ecol = m.group("end_col").upper()
    from openpyxl.utils.cell import column_index_from_string
    sc = column_index_from_string(scol); ec = column_index_from_string(ecol)
    return (srow, erow, sc, ec)

@dataclass
class EngineResult:
    warnings: List[str]
    details: Dict[str, Any]

class ExcelEngine:
    def open(self, path: str): ...
    def write_formula(self, sheet: str, a1_range: str, formula: str, fill_down: bool=False, named_range: Optional[str]=None) -> EngineResult: ...
    def create_pivot(self, source_sheet: str, target_sheet: str, rows: List[str], columns: List[str], values: List[str], filters: Dict[str, List[Any]], data_range: Optional[str]) -> EngineResult: ...
    def create_chart(self, target_sheet: str, chart_type: str, data_range: str, title: str, style: Dict[str, Any]) -> EngineResult: ...
    def save_as(self, path: str) -> None: ...
    def export_pdf(self, path_pdf: str) -> None: ...
    def close(self) -> None: ...

class FallbackEngine(ExcelEngine):
    def __init__(self, path: str):
        self.path = path
        if os.path.exists(path):
            self.wb = load_workbook(path)
        else:
            self.wb = Workbook()
        self.warnings: List[str] = []

    def open(self, path: str):
        return self

    def _ensure_sheet(self, name: str):
        if name in self.wb.sheetnames:
            return self.wb[name]
        return self.wb.create_sheet(title=name)

    def write_formula(self, sheet: str, a1_range: str, formula: str, fill_down: bool=False, named_range: Optional[str]=None) -> EngineResult:
        ws = self._ensure_sheet(sheet)
        srow, erow, sc, ec = a1_to_range(ws, a1_range if "!" not in a1_range else a1_range.split("!",1)[1])
        # 기록: 첫 셀에 formula, 필요시 fill_down
        from openpyxl.utils.cell import get_column_letter as _col
        for r in range(srow, erow+1):
            for c in range(sc, ec+1):
                cell = ws[f"{_col(c)}{r}"]
                cell.value = formula if (r == srow or fill_down) else cell.value
        if named_range:
            # 간단히 첫 셀만 이름정의 (범위 전체로 확장 가능)
            from openpyxl.workbook.defined_name import DefinedName
            first = f"{sheet}!{get_column_letter(sc)}{srow}:{get_column_letter(ec)}{erow}"
            dn = DefinedName(named_range, attr_text=first)
            self.wb.defined_names.append(dn)
        return EngineResult(self.warnings, {"range": a1_range, "formula": formula})

    def create_pivot(self, source_sheet: str, target_sheet: str, rows: List[str], columns: List[str], values: List[str], filters: Dict[str, List[Any]], data_range: Optional[str]) -> EngineResult:
        # pandas pivot_table로 결과표 생성
        df = pd.read_excel(self.path, sheet_name=source_sheet)
        if filters:
            for col, allowed in filters.items():
                if col in df.columns:
                    df = df[df[col].isin(allowed)]
        # 월 컬럼이 없다면 날짜에서 생성 시도
        if "월" in rows and "월" not in df.columns:
            # 날짜 후보 찾기
            date_col = None
            for c in df.columns:
                if "일" in c or "날짜" in c:
                    date_col = c; break
            if date_col:
                s = pd.to_datetime(df[date_col], errors="coerce")
                df["월"] = s.dt.to_period("M").astype(str)

        agg_map = {}
        value_cols: List[str] = []
        for spec in values:
            if ":" in spec:
                col, agg = spec.split(":",1)
            else:
                col, agg = spec, "sum"
            value_cols.append(col)
            agg_map[col] = agg

        if not rows and not columns:
            rows = [value_cols[0]]

        piv = pd.pivot_table(df, index=rows or None, columns=columns or None, values=value_cols, aggfunc=agg_map, fill_value=0, dropna=False)
        # 테이블을 일반 DataFrame으로
        piv = piv.reset_index()
        # 다중 컬럼일 경우 정리
        if isinstance(piv.columns, pd.MultiIndex):
            piv.columns = ["_".join([str(x) for x in tup if x != ""]) for tup in piv.columns.to_flat_index()]

        ws = self._ensure_sheet(target_sheet)
        # 기존 내용 삭제
        for row in ws["A1:ZZ1000"]:
            for cell in row:
                cell.value = None
        # 쓰기
        for j, col in enumerate(piv.columns, start=1):
            ws.cell(row=1, column=j, value=str(col))
        for i, (_, row) in enumerate(piv.iterrows(), start=2):
            for j, col in enumerate(piv.columns, start=1):
                ws.cell(row=i, column=j, value=row[col])
        return EngineResult(self.warnings, {"target_sheet": target_sheet, "rows": len(piv)})

    def create_chart(self, target_sheet: str, chart_type: str, data_range: str, title: str, style: Dict[str, Any]) -> EngineResult:
        # data_range 예: "피벗_매출!A1:D20"
        if "!" in data_range:
            sheet_name, rng = data_range.split("!",1)
        else:
            sheet_name, rng = target_sheet, data_range
        ws = self._ensure_sheet(sheet_name)
        srow, erow, sc, ec = a1_to_range(ws, rng)
        # 카테고리는 첫 열(A)로 가정, 데이터는 2..ec
        data = Reference(ws, min_col=sc+1, max_col=ec, min_row=srow, max_row=erow)
        cats = Reference(ws, min_col=sc, max_col=sc, min_row=srow+1, max_row=erow)

        if chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
            data = Reference(ws, min_col=sc+1, max_col=sc+1, min_row=srow, max_row=erow)
        else:
            chart = BarChart()

        chart.title = title
        chart.add_data(data, titles_from_data=True)
        if chart_type != "pie":
            chart.set_categories(cats)

        ws_target = self._ensure_sheet(target_sheet)
        ws_target.add_chart(chart, "A1")
        return EngineResult(self.warnings, {"chart_type": chart_type, "data_range": data_range})

    def save_as(self, path: str) -> None:
        self.wb.save(path)
        self.path = path

    def export_pdf(self, path_pdf: str) -> None:
        self.warnings.append("PDF export is not supported in FallbackEngine.")

    def close(self) -> None:
        pass

class ComEngine(ExcelEngine):
    # 구현은 개략적으로; 실제 실행은 Windows+Excel 환경 필요
    def __init__(self, path: str):
        if win32 is None:
            raise RuntimeError("win32com is not available")
        self.excel = win32.Dispatch("Excel.Application")
        self.excel.Visible = False
        self.wb = self.excel.Workbooks.Open(os.path.abspath(path))
        self.path = path
        self.warnings: List[str] = []

    def open(self, path: str):
        return self

    def _ws(self, name: str):
        try:
            return self.wb.Worksheets(name)
        except Exception:
            ws = self.wb.Worksheets.Add()
            ws.Name = name
            return ws

    def write_formula(self, sheet: str, a1_range: str, formula: str, fill_down: bool=False, named_range: Optional[str]=None) -> EngineResult:
        ws = self._ws(sheet)
        ws.Range(a1_range).Formula = formula
        if fill_down:
            ws.Range(a1_range).FillDown()
        if named_range:
            self.wb.Names.Add(Name=named_range, RefersTo=f"={sheet}!{a1_range}")
        return EngineResult(self.warnings, {"range": a1_range, "formula": formula})

    def create_pivot(self, source_sheet: str, target_sheet: str, rows: List[str], columns: List[str], values: List[str], filters: Dict[str, List[Any]], data_range: Optional[str]) -> EngineResult:
        ws_src = self._ws(source_sheet)
        rng = data_range or f"{source_sheet}!A1:Z9999"
        pcache = self.wb.PivotCaches().Create(SourceType=1, SourceData=rng)
        ws_tgt = self._ws(target_sheet)
        pt = pcache.CreatePivotTable(TableDestination=f"{target_sheet}!R1C1", TableName=f"PT_{int(time.time())}")
        # 필드 추가 (간단 버전)
        for r in rows:
            self.wb.ActiveSheet.PivotTables(pt.Name).PivotFields(r).Orientation = 1  # xlRowField
        for c in columns:
            self.wb.ActiveSheet.PivotTables(pt.Name).PivotFields(c).Orientation = 2  # xlColumnField
        for spec in values:
            col, agg = spec.split(":",1) if ":" in spec else (spec, "sum")
            pf = self.wb.ActiveSheet.PivotTables(pt.Name).PivotFields(col)
            self.wb.ActiveSheet.PivotTables(pt.Name).AddDataField(pf, f"{col} {agg}",  -4157 if agg=="sum" else -4106)  # xlSum / xlAverage
        return EngineResult(self.warnings, {"target_sheet": target_sheet})

    def create_chart(self, target_sheet: str, chart_type: str, data_range: str, title: str, style: Dict[str, Any]) -> EngineResult:
        ws = self._ws(target_sheet)
        ch = ws.ChartObjects().Add(10,10,500,300).Chart
        if chart_type == "line":
            ch.ChartType = 4   # xlLine
        elif chart_type == "pie":
            ch.ChartType = 5   # xlPie
        else:
            ch.ChartType = 51  # xlColumnClustered
        ch.SetSourceData(Source=data_range)
        ch.HasTitle = True
        ch.ChartTitle.Text = title
        return EngineResult(self.warnings, {"chart_type": chart_type})

    def save_as(self, path: str) -> None:
        self.wb.SaveAs(os.path.abspath(path))

    def export_pdf(self, path_pdf: str) -> None:
        self.wb.ExportAsFixedFormat(0, os.path.abspath(path_pdf))  # xlTypePDF

    def close(self) -> None:
        try:
            self.wb.Close(SaveChanges=True)
        finally:
            self.excel.Quit()

def pick_engine(path: str) -> ExcelEngine:
    # 우선 COM 시도(Windows + Excel), 실패시 Fallback
    if sys.platform == "win32" and win32 is not None:
        try:
            return ComEngine(path)
        except Exception as e:
            # fall back
            pass
    return FallbackEngine(path)
