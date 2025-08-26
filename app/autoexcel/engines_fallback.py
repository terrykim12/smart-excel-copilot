import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import re

# 스네이크/일치 매핑 유틸 (dedupe의 것을 재사용 중이라면 import로 대체 가능)
def _to_snake(name: str) -> str:
    if name is None: return name
    s = str(name).strip()
    s = re.sub(r"[^\w가-힣]+", "_", s, flags=re.UNICODE)
    s = re.sub(r"_{2,}", "_", s).strip("_")
    return s.lower()

def _colmap(df: pd.DataFrame):
    return { _to_snake(c): c for c in df.columns }

def _ensure_book(path):
    p = Path(path)
    if p.exists():
        return load_workbook(p)
    wb = Workbook()
    ws = wb.active; ws.title = "Sheet1"
    return wb

def write_formula(path, sheet, cell_range, formula, fill_down=False, named_range=None):
    wb = _ensure_book(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
    start_cell = ws[cell_range.split(":")[0]]
    ws[start_cell.coordinate] = formula
    if fill_down and ":" in cell_range:
        start, end = cell_range.split(":")
        end_row = int("".join(filter(str.isdigit, end)))
        col = "".join(filter(str.isalpha, start))
        for r in range(start_cell.row+1, end_row+1):
            ws[f"{col}{r}"] = f"={formula.lstrip('=')}"
    if named_range:
        wb.create_named_range(named_range, ws, cell_range)
    wb.save(path)

def _dedupe_columns(df: pd.DataFrame) -> pd.DataFrame:
    # 같은 이름의 열이 여러 개면 첫 번째만 남김 (피벗/그루퍼 안정화)
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated(keep="first")]
    return df

def _ensure_month_column(df: pd.DataFrame, colmap: dict) -> pd.DataFrame:
    # '월'이 이미 있고 DataFrame(중복 열)로 잡히면 첫 번째 열만 사용
    if '월' in df.columns and not isinstance(df['월'], pd.Series):
        df['월'] = df['월'].iloc[:, 0]
        return df
    # 없으면 날짜 후보에서 파생
    if '월' not in df.columns:
        for cand in ('주문일','order_date','date','날짜'):
            real = colmap.get(_to_snake(cand))
            if real and real in df.columns:
                dts = pd.to_datetime(df[real], errors="coerce")
                df = df.assign(월=dts.dt.to_period("M").astype(str))
                break
    return df

def create_pivot_from_df(df: pd.DataFrame, path, target_sheet, rows, values, filters=None):
    df = df.copy()

    # 0) 컬럼 중복 제거 (중복이면 groupby가 'not 1-dimensional' 에러)
    df = _dedupe_columns(df)
    colmap = _colmap(df)

    # 1) 필터 적용 (자연어 필터에서 넘어온 컬럼명 매핑)
    filters = filters or {}
    for col, allowed in list(filters.items()):
        real = colmap.get(_to_snake(col))
        if real and real in df.columns:
            df = df[df[real].isin(allowed)]

    # 2) rows/values 컬럼 매핑
    rows_real = []
    for r in rows or []:
        real = colmap.get(_to_snake(r))
        if real and real in df.columns:
            rows_real.append(real)

    agg = {}
    for (c, func) in values or []:
        real = colmap.get(_to_snake(c))
        if real and real in df.columns:
            agg[real] = func

    # 3) '월' 요청 시 파생 보강 (주문일/날짜에서 월 추출)
    want_month = any(_to_snake(r) == _to_snake('월') for r in (rows or []))
    if want_month:
        df = _ensure_month_column(df, colmap)
        if '월' in df.columns and '월' not in rows_real:
            rows_real = ['월'] + rows_real  # 월을 맨 앞에

    # 4) 실제 존재하는 컬럼만 최종 사용
    rows_real = [c for c in rows_real if c in df.columns]
    agg = {k:v for k,v in agg.items() if k in df.columns}
    if not agg:
        raise ValueError("피벗 값(value)으로 사용할 컬럼이 없습니다. (예: 금액:sum)")

    # 5) 피벗 생성
    pvt = pd.pivot_table(
        df,
        index=rows_real,
        values=list(agg.keys()),
        aggfunc=agg,
        fill_value=0,
        observed=False,
        dropna=False,
    ).reset_index()

    # 6) 엑셀로 쓰기
    wb = _ensure_book(path)
    ws = wb[target_sheet] if target_sheet in wb.sheetnames else wb.create_sheet(target_sheet)
    ws.delete_rows(1, ws.max_row)
    ws.append(list(pvt.columns))
    for row in pvt.itertuples(index=False):
        ws.append(list(row))
    wb.save(path)
    return pvt.shape

def add_chart(path, sheet, data_start_cell="A1", title="차트", chart_type="bar"):
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        # 피벗 시트 이름 자동 탐색 (유연성 ↑)
        cand = [s for s in wb.sheetnames if "피벗" in s]
        if not cand:
            raise ValueError(f"피벗 시트를 찾을 수 없습니다: {sheet}")
        sheet = cand[0]

    ws = wb[sheet]
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 2 or max_col < 2:
        raise ValueError(f"데이터가 부족합니다 (rows={max_row}, cols={max_col})")

    # 데이터/카테고리 범위 정의
    data = Reference(ws, min_col=2, max_col=max_col, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)

    # 차트 객체
    if chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    else:
        chart = BarChart()
    chart.title = title

    # 데이터 추가 (제목은 헤더 행에서)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # 일부 환경에서 series[].cat 이 개별 시리즈에 복제되지 않는 이슈 방지
    for s in getattr(chart, "series", []):
        if getattr(s, "cat", None) is None:
            s.cat = cats
        # values는 openpyxl 내부적으로 s.val 로 저장됩니다 (참고: s.values 아님)
        if getattr(s, "val", None) is None:
            # add_data에서 세팅이 안 됐다면 마지막 열을 기본값으로
            s.val = Reference(ws, min_col=max_col, max_col=max_col, min_row=1, max_row=max_row)

    # 배치 위치(데이터 오른쪽 2열 옆)
    anchor_col = get_column_letter(min(max_col + 2, 26))  # Z열 넘지 않게
    ws.add_chart(chart, f"{anchor_col}2")
    wb.save(path)
