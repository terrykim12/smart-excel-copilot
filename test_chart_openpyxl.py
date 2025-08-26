#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def test_chart_with_openpyxl():
    """openpyxl을 사용하여 차트 생성 기능을 테스트합니다."""
    
    # 테스트 파일 경로
    input_file = "data/automation/auto_demo.xlsx"
    output_file = "data/out/test_chart_openpyxl.xlsx"
    
    # 출력 디렉토리 생성
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    print("=== openpyxl 차트 생성 테스트 시작 ===")
    print(f"입력 파일: {input_file}")
    print(f"출력 파일: {output_file}")
    
    try:
        # 워크북 로드
        print("워크북 로드 중...")
        wb = load_workbook(input_file)
        
        # 첫 번째 시트 선택
        ws = wb.active
        print(f"활성 시트: {ws.title}")
        
        # 데이터 확인
        print(f"A1 값: {ws['A1'].value}")
        print(f"A2 값: {ws['A2'].value}")
        print(f"B1 값: {ws['B1'].value}")
        print(f"B2 값: {ws['B2'].value}")
        
        # 데이터 범위 찾기
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"데이터 범위: {max_row}행 x {max_col}열")
        
        # 차트 시트 생성
        chart_sheet = wb.create_sheet("차트")
        
        # 피벗 테이블과 유사한 데이터 요약 (간단한 버전)
        # 날짜별 금액 합계 계산
        date_amounts = {}
        for row in range(2, max_row + 1):  # 헤더 제외
            date_val = ws.cell(row=row, column=1).value  # 날짜 (열 1)
            amount_val = ws.cell(row=row, column=4).value  # 금액 (열 4)
            
            if date_val and amount_val:
                if isinstance(amount_val, str):
                    try:
                        amount_val = float(amount_val.replace(',', ''))
                    except:
                        continue
                
                # 날짜를 문자열로 변환
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)
                
                if date_str in date_amounts:
                    date_amounts[date_str] += amount_val
                else:
                    date_amounts[date_str] = amount_val
        
        print(f"날짜별 금액 합계: {date_amounts}")
        
        # 차트 데이터를 차트 시트에 작성
        chart_sheet['A1'] = '날짜'
        chart_sheet['B1'] = '금액합계'
        
        row = 2
        for date, amount in date_amounts.items():
            chart_sheet[f'A{row}'] = date
            chart_sheet[f'B{row}'] = amount
            row += 1
        
        # 차트 생성
        chart = BarChart()
        chart.title = "날짜별 금액 합계"
        chart.x_axis.title = "날짜"
        chart.y_axis.title = "금액"
        
        # 데이터 범위 설정
        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(date_amounts)+1)
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(date_amounts)+1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # 차트를 차트 시트에 추가
        chart_sheet.add_chart(chart, "D2")
        
        # 파일 저장
        wb.save(output_file)
        print(f"✅ 차트 생성 성공! 출력 파일: {output_file}")
        
        # 파일 존재 확인
        if os.path.exists(output_file):
            print(f"✅ 출력 파일이 정상적으로 생성되었습니다.")
            file_size = os.path.getsize(output_file)
            print(f"파일 크기: {file_size:,} bytes")
        else:
            print("❌ 출력 파일이 생성되지 않았습니다.")
            
    except Exception as e:
        print(f"❌ 차트 생성 실패: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True

if __name__ == "__main__":
    success = test_chart_with_openpyxl()
    if success:
        print("\n🎉 모든 테스트가 성공적으로 완료되었습니다!")
    else:
        print("\n💥 테스트 중 오류가 발생했습니다.")
        sys.exit(1)
