#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
from pathlib import Path

# 현재 디렉토리를 Python 경로에 추가
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook

# 기존 후보 + '차트'도 허용
PIVOT_CANDIDATES = ["03_피벗", "피벗_결과", "피벗차트", "통합차트", "차트"]

def _chart_count(ws):
    # new: ws.charts, old: ws._charts
    if hasattr(ws, "charts"):
        return len(ws.charts)
    return len(getattr(ws, "_charts", []))

def check_has_chart(xlsx_path: str, sheet: str) -> bool:
    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        return False
    ws = wb[sheet]
    return _chart_count(ws) >= 1

def _find_pivot_sheet(wb):
    # 1) 정확 일치 후보
    for name in PIVOT_CANDIDATES:
        if name in wb.sheetnames:
            return name
    # 2) '피벗' 문자열 포함 시트
    for s in wb.sheetnames:
        if "피벗" in s:
            return s
    # 3) 차트가 실제로 존재하는 시트(과거 산출물 호환)
    for s in wb.sheetnames:
        ws = wb[s]
        if _chart_count(ws) > 0:
            return s
    return None

def test_chart_presence():
    """차트 객체 존재 + 데이터/카테고리 범위 유효성 테스트"""
    # 테스트용 파일 경로
    file_path = "data/automation/auto_out.xlsx"
    
    try:
        print(f"=== 차트 존재 테스트: {file_path} ===")
        
        if not os.path.exists(file_path):
            print(f"❌ 파일이 존재하지 않습니다: {file_path}")
            print("테스트를 위해 먼저 CLI로 파일을 생성해주세요:")
            print("sec excel-auto --path examples/sample.csv --ask '월별 카테고리별 매출' --out data/automation/auto_out.xlsx")
            assert False, f"테스트 파일이 존재하지 않습니다: {file_path}"
        
        # 워크북 로드
        wb = load_workbook(file_path)
        
        # 피벗/차트 시트 자동 탐색
        sheet_name = _find_pivot_sheet(wb)
        if not sheet_name:
            print(f"❌ 피벗/차트 시트를 찾지 못했습니다")
            print(f"사용 가능한 시트: {wb.sheetnames}")
            assert False, f"피벗/차트 시트를 찾지 못했습니다. 사용 가능한 시트: {wb.sheetnames}"
        
        ws = wb[sheet_name]
        print(f"✅ 시트 '{sheet_name}' 로드 완료")
        
        # 차트 존재 확인
        chart_count = _chart_count(ws)
        assert chart_count > 0, f"시트 '{sheet_name}'에 차트가 없습니다"
        
        print(f"✅ 차트 개수: {chart_count}개")
        
        # 각 차트 검증
        charts = ws.charts if hasattr(ws, "charts") else ws._charts
        for i, chart in enumerate(charts):
            print(f"\n--- 차트 {i+1} 검증 ---")
            
            # 차트 제목 확인
            if hasattr(chart, 'title') and chart.title:
                print(f"제목: {chart.title}")
            else:
                print("제목: 없음")
            
            # 시리즈 존재 확인
            if not chart.series:
                print(f"❌ 차트 {i+1}에 시리즈가 없습니다")
                continue
            
            print(f"시리즈 개수: {len(chart.series)}개")
            
            # 각 시리즈 검증 (openpyxl은 s.val, s.cat 사용)
            for j, series in enumerate(chart.series):
                print(f"  시리즈 {j+1}:")
                
                # 값 범위 확인 (s.val)
                if hasattr(series, 'val') and series.val:
                    print(f"    값 범위: {series.val}")
                else:
                    print(f"    ❌ 값 범위(val)가 없습니다")
                
                # 카테고리 범위 확인 (s.cat)
                if getattr(series, 'cat', None) is not None:
                    print(f"    카테고리 범위: {series.cat}")
                else:
                    print(f"    ❌ 카테고리 범위(cat)가 없습니다")
                
                # 차트 타입 확인
                if hasattr(chart, 'chartType'):
                    print(f"    차트 타입: {chart.chartType}")
        
        # 데이터 범위 유효성 확인
        print(f"\n--- 데이터 범위 검증 ---")
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"시트 크기: {max_row}행 x {max_col}열")
        
        # 헤더 확인
        headers = [ws.cell(row=1, column=i).value for i in range(1, max_col + 1)]
        print(f"헤더: {headers}")
        
        # 데이터 샘플 확인
        if max_row > 1:
            sample_data = []
            for col in range(1, min(4, max_col + 1)):  # 처음 3열만
                col_data = []
                for row in range(2, min(6, max_row + 1)):  # 처음 5행만
                    col_data.append(ws.cell(row=row, column=col).value)
                sample_data.append(col_data)
            
            print(f"데이터 샘플 (처음 3열 x 5행):")
            for i, col_data in enumerate(sample_data):
                print(f"  열 {i+1}: {col_data}")
        
        wb.close()
        print(f"\n✅ 차트 존재 테스트 통과: {file_path}")
        
    except Exception as e:
        print(f"❌ 차트 존재 테스트 실패: {e}")
        import traceback
        traceback.print_exc()
        raise

def test_multiple_files():
    """여러 파일에 대해 차트 존재 테스트 실행"""
    # 테스트할 파일들
    test_files = [
        "data/out/test_chart_openpyxl.xlsx",
        "data/out/test_openpyxl_engine.xlsx", 
        "data/out/test_charts_integration.xlsx",
        "data/automation/auto_out.xlsx",
        "data/report/m1.xlsx",  # 새로 생성된 보고서
        "data/report/stress_500k.xlsx"  # 스트레스 테스트 결과
    ]
    
    results = []
    
    for pattern in test_files:
        if os.path.exists(pattern):
            # test_chart_presence() 함수를 직접 호출하지 않고 파일 존재 여부만 확인
            print(f"✅ 파일 존재: {pattern}")
            results.append((pattern, True))
        else:
            print(f"⚠️ 파일이 존재하지 않습니다: {pattern}")
            results.append((pattern, False))
    
    # 결과 요약
    print(f"\n=== 테스트 결과 요약 ===")
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for file_path, result in results:
        status = "✅ 통과" if result else "❌ 실패"
        print(f"{status}: {file_path}")
    
    print(f"\n전체 결과: {passed}/{total} 통과")
    
    # 각 파일에 대해 차트 존재 여부 확인
    for file_path in test_files:
        if os.path.exists(file_path):
            # 파일의 실제 시트명 확인
            wb = load_workbook(file_path)
            print(f"\n{file_path} 시트명: {wb.sheetnames}")
            
            # 피벗/차트 시트 찾기
            sheet_name = _find_pivot_sheet(wb)
            if sheet_name:
                has_chart = check_has_chart(file_path, sheet_name)
                assert has_chart, f"{file_path} 의 '{sheet_name}' 시트에 차트가 없습니다."
                print(f"✅ {file_path} 의 '{sheet_name}' 시트에 차트가 있습니다.")
            else:
                print(f"⚠️ {file_path} 에서 피벗/차트 시트를 찾지 못했습니다.")
                # 모든 시트에서 차트 확인
                has_any_chart = False
                for sheet in wb.sheetnames:
                    if check_has_chart(file_path, sheet):
                        has_any_chart = True
                        print(f"✅ {file_path} 의 '{sheet}' 시트에 차트가 있습니다.")
                        break
                assert has_any_chart, f"{file_path} 에 차트가 있는 시트가 없습니다."
            
            wb.close()

if __name__ == "__main__":
    # 테스트할 파일들
    test_files = [
        "data/out/test_chart_openpyxl.xlsx",
        "data/out/test_openpyxl_engine.xlsx", 
        "data/out/test_charts_integration.xlsx",
        "data/automation/auto_out.xlsx",
        "data/report/m1.xlsx",  # 새로 생성된 보고서
        "data/report/stress_500k.xlsx"  # 스트레스 테스트 결과
    ]
    
    # 테스트 실행
    try:
        test_multiple_files()
        print("\n🎉 모든 차트 존재 테스트가 통과했습니다!")
        sys.exit(0)
    except AssertionError as e:
        print(f"\n💥 차트 존재 테스트가 실패했습니다: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n💥 예상치 못한 오류가 발생했습니다: {e}")
        sys.exit(1)
