#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys

# 현재 디렉토리를 Python 경로에 추가
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'app'))

from autoexcel.charts import run_pivot_chart_openpyxl

def test_charts_integration():
    """charts.py의 openpyxl 통합 기능을 테스트합니다."""
    
    # 테스트 파일 경로
    input_file = "data/automation/auto_demo.xlsx"
    output_file = "data/out/test_charts_integration.xlsx"
    
    print("=== charts.py 통합 테스트 시작 ===")
    print(f"입력 파일: {input_file}")
    print(f"출력 파일: {output_file}")
    
    try:
        # 피벗 차트 생성 테스트
        success = run_pivot_chart_openpyxl(
            path_in=input_file,
            path_out=output_file,
            src_sheet="원본",
            tgt_sheet="통합차트",
            rows=["카테고리"],  # 카테고리별로 그룹화
            values=["금액"],    # 금액 합계
            chart_type="bar"    # 막대 차트
        )
        
        if success:
            print(f"✅ 통합 차트 생성 성공! 출력 파일: {output_file}")
            
            # 파일 존재 확인
            if os.path.exists(output_file):
                print(f"✅ 출력 파일이 정상적으로 생성되었습니다.")
                file_size = os.path.getsize(output_file)
                print(f"파일 크기: {file_size:,} bytes")
                
                # 파일 내용 간단 확인
                from openpyxl import load_workbook
                wb = load_workbook(output_file)
                if "통합차트" in wb.sheetnames:
                    ws = wb["통합차트"]
                    print(f"✅ 차트 시트 생성 확인: {ws.title}")
                    print(f"시트 내용: A1={ws['A1'].value}, A2={ws['A2'].value}")
                wb.close()
            else:
                print("❌ 출력 파일이 생성되지 않았습니다.")
        else:
            print("❌ 통합 차트 생성 실패")
            return False
            
    except Exception as e:
        print(f"❌ 테스트 실패: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True

if __name__ == "__main__":
    success = test_charts_integration()
    if success:
        print("\n🎉 모든 테스트가 성공적으로 완료되었습니다!")
    else:
        print("\n💥 테스트 중 오류가 발생했습니다.")
        sys.exit(1)
